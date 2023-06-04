from datetime import datetime, date

from django.core.management.base import BaseCommand, CommandError
from openpyxl.reader.excel import load_workbook

from apps.timeregistrations.models import TimeRegistration


class Command(BaseCommand):
    help = "Import excel sheet with hour data"

    def handle(self, *args, **options):
        workbook = load_workbook("/var/backups/Uren.xlsx")
        projects = set()
        for worksheet_name in workbook.sheetnames:
            worksheet = workbook[worksheet_name]
            worksheets_without_project = ["November 2020", "December 2020", "Januari 2021", "Februari 2021"]
            if worksheet_name in worksheets_without_project:
                external_ref_row_number = 2
                start_time_row_number = 3
                end_time_row_number = 4
                project_row_number = None
            else:
                project_row_number = 2
                external_ref_row_number = 3
                start_time_row_number = 4
                end_time_row_number = 5
            for index, row in enumerate(worksheet.iter_rows(), start=1):
                if (
                        row[0].value
                        and row[start_time_row_number].value
                        and row[end_time_row_number].value
                        and isinstance(row[0].value, datetime)
                ):
                    date = row[0].value
                    description = row[1].value
                    project = row[project_row_number].value if project_row_number else None
                    if project:
                        projects.add(project)

                    external_reference = row[external_ref_row_number].value
                    start_time = row[start_time_row_number].value
                    end_time = row[end_time_row_number].value
                    start = None
                    end = None

                    try:
                        start = datetime.combine(date.date(), start_time)
                    except TypeError:
                        print("ERROR START: ", worksheet_name, description, index, start_time)
                    try:
                        end = datetime.combine(date.date(), end_time)
                    except TypeError:
                        print("ERROR END: ", worksheet_name, description, index, end_time)

                    # print(f"{description} ({external_reference}) for project {project}: {start} - {end}")
                    mapping = {
                        "JURSOFT": "GO_01",
                        "ALGEMEEN": "GO_03",
                        "BRUNEL": "GO_02",
                        "CASER": "GO_01",
                        "Algemeen": "GO_03",
                        "CRM": "GO_04",
                        "AZL": "GO_05",
                        None: "GO_01"
                    }
                    time_registration, created = TimeRegistration.objects.get_or_create(
                        start=start,
                        end=end,
                        defaults={
                            "project_id": mapping.get(project),
                            "user_id": 1,
                            "description": description,
                            "external_reference": external_reference
                        }
                    )

